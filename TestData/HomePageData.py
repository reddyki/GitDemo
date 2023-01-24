
import openpyxl


class HomePageData:

    test_HomePage_data = [{"firstname":"rahul", "lastname":"shetty", "password":"1234", "gender":"Male"}, {"firstname":"Anshika", "lastname":"shetty", "password":"4567", "gender":"Female"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\kisho\\OneDrive\\Documents\\PyexcelDemo.xlsx")
        sheet = book.active

        for i in range(1, sheet.max_row + 1):  # to get row
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):  # to get column
                    # Dict["firstname"] = "rahul"
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]